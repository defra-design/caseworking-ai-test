"""
Generate data/grasslands-cases.xlsx from data/grasslands-cases.js.

This is the "export" half of the round-trip. Run it whenever the .js data
file changes and you want a fresh spreadsheet to edit. The "import" half
(import-xlsx.py) writes the .js back from the edited spreadsheet.

Usage:  python tools/grasslands-caselist/build-xlsx.py
Requires: node (to read the .js data) and openpyxl (pip install openpyxl).
"""
import json
import subprocess
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

# Fixed month list so date parsing/formatting is locale-independent.
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
DATE_FORMAT = "d mmm yyyy"  # Excel display format → "6 Jan 2025"


def parse_date(value):
    """Parse a 'D MMM YYYY' string (e.g. '6 Jan 2025') to a datetime, or None."""
    try:
        day, mon, year = str(value).strip().split()
        return datetime(int(year), MONTHS.index(mon.title()) + 1, int(day))
    except (ValueError, IndexError):
        return None

APP = Path(__file__).resolve().parents[2]          # .../app
JS = APP / "data" / "grasslands-cases.js"
XLSX = APP / "data" / "grasslands-cases.xlsx"

# Column order MUST match the object key order written back by import-xlsx.py.
COLUMNS = ["id", "business", "sbi", "submitted", "status", "tag", "team", "assignee"]

# Allowed values — keep these in step with the DATA RULES comment in the .js.
STATUS_TO_TAG = [
    ("Application received", "grey"),
    ("In review", "blue"),
    ("On hold", "yellow"),
    ("Agreement drafted", "blue"),
    ("Agreement offered", "blue"),
    ("Agreement accepted", "green"),
    ("Rejected", "red"),
    ("Withdrawn", "orange"),
    ("live", ""),
]
STATUSES = [s for s, _ in STATUS_TO_TAG]
TAGS = ["grey", "blue", "yellow", "green", "red", "orange"]
TEAMS = ["A", "B", "C"]


def load_cases():
    """Read the cases array out of the .js file via node, as JSON."""
    out = subprocess.run(
        ["node", "-e", f"process.stdout.write(JSON.stringify(require({json.dumps(str(JS))}).cases))"],
        capture_output=True, text=True,
    )
    if out.returncode != 0:
        sys.exit("Failed to read grasslands-cases.js via node:\n" + out.stderr)
    return json.loads(out.stdout)


def main():
    cases = load_cases()
    wb = openpyxl.Workbook()

    # --- Cases sheet ---
    ws = wb.active
    ws.title = "Cases"
    header_fill = PatternFill("solid", fgColor="00703C")  # GOV.UK green
    header_font = Font(bold=True, color="FFFFFF")
    for c, name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    for r, case in enumerate(cases, start=2):
        for c, key in enumerate(COLUMNS, start=1):
            value = case.get(key, "")
            if key == "submitted":
                dt = parse_date(value)
                if dt is not None:
                    cell = ws.cell(row=r, column=c, value=dt)
                    cell.number_format = DATE_FORMAT  # real date → sorts chronologically
                    continue
            cell = ws.cell(row=r, column=c, value=value)
            cell.number_format = "@"  # force TEXT so sbi/id keep leading zeros

    # Sensible widths
    widths = {"id": 10, "business": 32, "sbi": 13, "submitted": 14,
              "status": 20, "tag": 9, "team": 7, "assignee": 14}
    for c, key in enumerate(COLUMNS, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = widths[key]
    ws.freeze_panes = "A2"

    # Dropdowns so edits stay valid (Excel only; ignored elsewhere)
    last = len(cases) + 1
    def add_dv(col_key, values):
        col = openpyxl.utils.get_column_letter(COLUMNS.index(col_key) + 1)
        dv = DataValidation(type="list", formula1='"' + ",".join(values) + '"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}{last}")
    add_dv("status", STATUSES)
    add_dv("tag", TAGS)
    add_dv("team", TEAMS)

    # --- Reference sheet ---
    ref = wb.create_sheet("Reference")
    ref["A1"] = "Status → tag mapping (use these only)"
    ref["A1"].font = Font(bold=True)
    ref["A2"], ref["B2"] = "status", "tag"
    ref["A2"].font = ref["B2"].font = Font(bold=True)
    for i, (status, tag) in enumerate(STATUS_TO_TAG, start=3):
        ref.cell(row=i, column=1, value=status)
        ref.cell(row=i, column=2, value=tag)
    note_row = len(STATUS_TO_TAG) + 4
    notes = [
        "Notes:",
        "- sbi is a unique 9-digit string (leading zeros matter — cells are TEXT).",
        "- team is A, B or C; leave blank with assignee 'Not assigned' for unassigned cases.",
        "- 'submitted' is a real date (shown d mmm yyyy) so the column sorts chronologically.",
        "- 'live' (Golden Grange) has a blank tag.",
        "- Do not rename the columns or reorder the sheet — import-xlsx.py reads them by name.",
    ]
    for i, line in enumerate(notes):
        ref.cell(row=note_row + i, column=1, value=line)
    ref.column_dimensions["A"].width = 60
    ref.column_dimensions["B"].width = 12

    wb.save(XLSX)
    print(f"Wrote {XLSX.relative_to(APP)} ({len(cases)} cases).")


if __name__ == "__main__":
    main()
