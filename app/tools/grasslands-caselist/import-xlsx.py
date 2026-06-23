"""
Write data/grasslands-cases.js from an edited data/grasslands-cases.xlsx.

This is the "re-upload" half of the round-trip: edit the spreadsheet, then run
this to regenerate the .js data file the Grasslands caselist reads. The leading
comment block (DATA RULES) in the existing .js is preserved.

Usage:  python tools/grasslands-caselist/import-xlsx.py [path-to-xlsx]
        (defaults to data/grasslands-cases.xlsx)
Requires: openpyxl (pip install openpyxl).
"""
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

APP = Path(__file__).resolve().parents[2]          # .../app
JS = APP / "data" / "grasslands-cases.js"
DEFAULT_XLSX = APP / "data" / "grasslands-cases.xlsx"

COLUMNS = ["id", "business", "sbi", "submitted", "status", "tag", "team", "assignee"]

# Fixed month list so date formatting is locale-independent.
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def fmt_date(value):
    """A real date cell -> 'D MMM YYYY' (e.g. '6 Jan 2025'); otherwise pass through as text."""
    if isinstance(value, (datetime, date)):
        return f"{value.day} {MONTHS[value.month - 1]} {value.year}"
    return "" if value is None else str(value).strip()


def js_string(value):
    """Render a Python value as a double-quoted JS string literal."""
    s = "" if value is None else str(value)
    s = s.replace("\\", "\\\\").replace('"', '\\"')
    return '"' + s + '"'


def header_block():
    """Reuse the comment header from the existing .js (everything before module.exports)."""
    if JS.exists():
        head = []
        for line in JS.read_text(encoding="utf-8").splitlines():
            if line.startswith("module.exports"):
                break
            head.append(line)
        if head:
            return "\n".join(head).rstrip() + "\n"
    return "// Data reference file: the full Grasslands caselist case set.\n"


def read_rows(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "Cases" not in wb.sheetnames:
        sys.exit('No "Cases" sheet found in the spreadsheet.')
    ws = wb["Cases"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        sys.exit("The Cases sheet is empty.")

    header = [("" if h is None else str(h).strip()) for h in rows[0]]
    missing = [c for c in COLUMNS if c not in header]
    if missing:
        sys.exit("Missing column(s) in the Cases sheet: " + ", ".join(missing))
    idx = {c: header.index(c) for c in COLUMNS}

    cases = []
    for r in rows[1:]:
        # skip fully-blank rows
        if all(v is None or str(v).strip() == "" for v in r):
            continue
        case = {}
        for key in COLUMNS:
            v = r[idx[key]]
            if key == "submitted":
                case[key] = fmt_date(v)
            else:
                case[key] = "" if v is None else str(v).strip()
        if not case["id"]:
            continue
        cases.append(case)
    return cases


def main():
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx_path.exists():
        sys.exit(f"Spreadsheet not found: {xlsx_path}")

    cases = read_rows(xlsx_path)

    lines = []
    for case in cases:
        pairs = ", ".join(f"{key}: {js_string(case[key])}" for key in COLUMNS)
        lines.append("  { " + pairs + " }")

    body = "module.exports = {\n  cases: [\n" + ",\n".join(lines) + "\n  ]\n}\n"
    # Capture the comment header BEFORE opening the file for write (which truncates it).
    content = header_block() + body
    # Write LF line endings (match the existing data file; avoids noisy CRLF diffs).
    with JS.open("w", encoding="utf-8", newline="\n") as fh:
        fh.write(content)
    print(f"Wrote {JS.relative_to(APP)} ({len(cases)} cases) from {xlsx_path.name}.")


if __name__ == "__main__":
    main()
